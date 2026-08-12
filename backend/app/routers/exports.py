import io
import calendar
from datetime import date

from fastapi import APIRouter, Depends, Request
from fastapi.responses import StreamingResponse, HTMLResponse
from fastapi.templating import Jinja2Templates
from sqlmodel import Session, select

from app.database import get_session
from app.models.invoices import Invoice
from app.models.parties import Party, OldGoldExchange
from app.models.expenses import Expense, ExpenseCategory
from app.models.inventory import StockLedger
from app.models.products import Product
from app.models.payments import PaymentEvent
from app.models.shop import FinancialYear

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

router    = APIRouter(prefix="/exports", tags=["Exports"])
templates = Jinja2Templates(directory="app/templates")


# ── STYLE HELPERS ─────────────────────────────────────────────────────────────

HEADER_FILL   = PatternFill("solid", fgColor="1A237E")   
HEADER_FONT   = Font(color="FFFFFF", bold=True, size=10)
SUBHEAD_FILL  = PatternFill("solid", fgColor="E8EAF6")   
BOLD_FONT     = Font(bold=True, size=10)
NORMAL_FONT   = Font(size=10)
RUPEE_FORMAT  = '#,##0.00'
THIN_BORDER   = Border(
    left   = Side(style="thin"),
    right  = Side(style="thin"),
    top    = Side(style="thin"),
    bottom = Side(style="thin"),
)

def _header_row(ws, cols: list, row: int = 1):
    """Write a styled header row. cols = list of (header_text, width)."""
    for col_idx, (text, width) in enumerate(cols, start=1):
        cell = ws.cell(row=row, column=col_idx, value=text)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

def _data_row(ws, values: list, row: int, number_cols: set = None):
    """Write a data row with borders. number_cols = set of 1-based col indices that are amounts."""
    number_cols = number_cols or set()
    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col_idx, value=value)
        cell.font   = NORMAL_FONT
        cell.border = THIN_BORDER
        if col_idx in number_cols and isinstance(value, (int, float)):
            cell.number_format = RUPEE_FORMAT
            cell.alignment     = Alignment(horizontal="right")

def _total_row(ws, values: list, row: int, number_cols: set = None):
    """Write a bold totals row."""
    number_cols = number_cols or set()
    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col_idx, value=value)
        cell.font   = BOLD_FONT
        cell.fill   = SUBHEAD_FILL
        cell.border = THIN_BORDER
        if col_idx in number_cols and isinstance(value, (int, float)):
            cell.number_format = RUPEE_FORMAT
            cell.alignment     = Alignment(horizontal="right")

def _stream(wb: openpyxl.Workbook, filename: str) -> StreamingResponse:
    """Save workbook to bytes and return as download response."""
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


# ── INDEX PAGE ────────────────────────────────────────────────────────────────

@router.get("/", response_class=HTMLResponse)
def exports_index(request: Request):
    return templates.TemplateResponse(
        request=request,
        name="exports/index.html",
        context={},
    )


# ── 1. SALES REGISTER ─────────────────────────────────────────────────────────

@router.get("/sales-register")
def export_sales_register(
    month: str = "",
    fy_id: int = 0,
    session: Session = Depends(get_session),
):
    stmt = (
        select(Invoice)
        .where(Invoice.invoice_type == "sale")
        .where(Invoice.is_cancelled == False)
        .order_by(Invoice.invoice_date)
    )
    selected_fy = None
    if fy_id:
        selected_fy = session.get(FinancialYear, fy_id)
    if not selected_fy:
        selected_fy = session.exec(
            select(FinancialYear).where(FinancialYear.is_active == True)
        ).first()

    if selected_fy:
        stmt = stmt.where(Invoice.invoice_date >= selected_fy.start_date)
        stmt = stmt.where(Invoice.invoice_date <= selected_fy.end_date)

    if month:
        try:
            y, m = int(month[:4]), int(month[5:7])
            last = calendar.monthrange(y, m)[1]
            stmt = stmt.where(Invoice.invoice_date >= date(y, m, 1))
            stmt = stmt.where(Invoice.invoice_date <= date(y, m, last))
        except Exception:
            pass

    invoices = session.exec(stmt).all()

    party_ids = {inv.party_id for inv in invoices if inv.party_id}
    parties   = {p.id: p for p in session.exec(
        select(Party).where(Party.id.in_(party_ids))
    ).all()} if party_ids else {}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales Register"
    ws.row_dimensions[1].height = 30

    cols = [
        ("Invoice No.",    16), ("Date",          12), ("Party Name",     22),
        ("GSTIN",          18), ("Bill Type",      10), ("Subtotal (₹)",   14),
        ("CGST (₹)",       12), ("SGST (₹)",       12), ("Making (₹)",     12),
        ("Making CGST (₹)",14), ("Making SGST (₹)",14), ("Old Gold (₹)",   12),
        ("Discount (₹)",   12), ("Round Off (₹)",  12), ("Grand Total (₹)",16),
        ("Amount Paid (₹)",16), ("Amount Due (₹)", 16), ("Payment Status", 14),
        ("GST Status",     14),
    ]
    _header_row(ws, cols, row=1)

    AMT_COLS = {6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17}
    totals   = {c: 0.0 for c in AMT_COLS}

    for r_idx, inv in enumerate(invoices, start=2):
        party = parties.get(inv.party_id)
        row = [
            inv.invoice_number,
            inv.invoice_date.strftime("%d-%m-%Y"),
            party.name if party else "Walk-in",
            inv.party_gstin or (party.gstin if party else "") or "",
            inv.bill_category.upper(),
            inv.subtotal,
            inv.total_cgst,
            inv.total_sgst,
            inv.total_making_charges or 0,
            inv.making_cgst or 0,
            inv.making_sgst or 0,
            inv.old_gold_value or 0,
            inv.discount or 0,
            inv.round_off or 0,
            inv.grand_total,
            inv.amount_paid,
            inv.amount_due,
            inv.payment_status.upper(),
            inv.gst_status.replace("_", " ").upper(),
        ]
        _data_row(ws, row, r_idx, AMT_COLS)
        for c in AMT_COLS:
            totals[c] += row[c - 1] if isinstance(row[c - 1], (int, float)) else 0

    t_row = len(invoices) + 2
    total_values = ["TOTAL", f"{len(invoices)} bills", "", "", ""]
    for c_idx in range(6, 20):
        total_values.append(round(totals.get(c_idx, 0), 2))
    _total_row(ws, total_values, t_row, AMT_COLS)

    fname = f"sales_register_{selected_fy.label if selected_fy else 'all'}_{month or 'all'}.xlsx"
    return _stream(wb, fname)


# ── 2. PURCHASE REGISTER ──────────────────────────────────────────────────────

@router.get("/purchase-register")
def export_purchase_register(
    month: str = "",
    fy_id: int = 0,
    session: Session = Depends(get_session),
):
    stmt = (
        select(Invoice)
        .where(Invoice.invoice_type == "purchase")
        .where(Invoice.is_cancelled == False)
        .order_by(Invoice.invoice_date)
    )

    selected_fy = None
    if fy_id:
        selected_fy = session.get(FinancialYear, fy_id)
    if not selected_fy:
        selected_fy = session.exec(
            select(FinancialYear).where(FinancialYear.is_active == True)
        ).first()

    if selected_fy:
        stmt = stmt.where(Invoice.invoice_date >= selected_fy.start_date)
        stmt = stmt.where(Invoice.invoice_date <= selected_fy.end_date)

    if month:
        try:
            y, m = int(month[:4]), int(month[5:7])
            last = calendar.monthrange(y, m)[1]
            stmt = stmt.where(Invoice.invoice_date >= date(y, m, 1))
            stmt = stmt.where(Invoice.invoice_date <= date(y, m, last))
        except Exception:
            pass

    invoices = session.exec(stmt).all()

    party_ids = {inv.party_id for inv in invoices if inv.party_id}
    parties   = {p.id: p for p in session.exec(
        select(Party).where(Party.id.in_(party_ids))
    ).all()} if party_ids else {}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Purchase Register"
    ws.row_dimensions[1].height = 30

    cols = [
        ("Invoice No.",    16), ("Date",          12), ("Supplier Name",  22),
        ("GSTIN",          18), ("Subtotal (₹)",   14), ("CGST (₹)",       12),
        ("SGST (₹)",       12), ("Grand Total (₹)",16), ("Amount Paid (₹)",16),
        ("Amount Due (₹)", 16), ("Payment Status", 14), ("GST Status",     14),
    ]
    _header_row(ws, cols, row=1)

    AMT_COLS = {5, 6, 7, 8, 9, 10}
    totals   = {c: 0.0 for c in AMT_COLS}

    for r_idx, inv in enumerate(invoices, start=2):
        party = parties.get(inv.party_id)
        row = [
            inv.invoice_number,
            inv.invoice_date.strftime("%d-%m-%Y"),
            party.name if party else "—",
            party.gstin if party else "",
            inv.subtotal,
            inv.total_cgst,
            inv.total_sgst,
            inv.grand_total,
            inv.amount_paid,
            inv.amount_due,
            inv.payment_status.upper(),
            inv.gst_status.replace("_", " ").upper(),
        ]
        _data_row(ws, row, r_idx, AMT_COLS)
        for c in AMT_COLS:
            totals[c] += row[c - 1] if isinstance(row[c - 1], (int, float)) else 0

    t_row = len(invoices) + 2
    total_values = ["TOTAL", f"{len(invoices)} bills", "", ""]
    for c_idx in range(5, 13):
        total_values.append(round(totals.get(c_idx, 0), 2))
    _total_row(ws, total_values, t_row, AMT_COLS)

    fname = f"purchase_register_{selected_fy.label if selected_fy else 'all'}_{month or 'all'}.xlsx"
    return _stream(wb, fname)


# ── 3. PARTY LEDGER ───────────────────────────────────────────────────────────

@router.get("/party-ledger/{party_id}")
def export_party_ledger(
    party_id: int,
    session: Session = Depends(get_session),
):
    party = session.get(Party, party_id)
    if not party:
        from fastapi import HTTPException
        raise HTTPException(status_code=404, detail="Party not found")

    invoices = session.exec(
        select(Invoice)
        .where(Invoice.party_id == party_id)
        .where(Invoice.is_cancelled == False)
        .order_by(Invoice.invoice_date)
    ).all()

    payments = session.exec(
        select(PaymentEvent)
        .where(PaymentEvent.party_id == party_id)
        .order_by(PaymentEvent.event_date)
    ).all()
    payment_map = {}
    for p in payments:
        payment_map.setdefault(p.invoice_id, []).append(p)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{party.name[:20]} Ledger"

    ws["A1"] = "Party Ledger"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = party.name
    ws["A2"].font = Font(bold=True, size=11)
    ws["A3"] = f"Phone: {party.phone or '—'}   GSTIN: {party.gstin or '—'}   Type: {party.type.upper()}"
    ws["A3"].font = Font(italic=True, size=9, color="888888")
    ws["A4"] = f"Report generated: {date.today().strftime('%d-%m-%Y')}"
    ws["A4"].font = Font(italic=True, size=9, color="888888")

    cols = [
        ("Invoice No.",    16), ("Date",          12), ("Type",           10),
        ("Grand Total (₹)",16), ("Amount Paid (₹)",16), ("Amount Due (₹)", 16),
        ("Payment Status", 14), ("GST Status",     14), ("Payments Made",  30),
    ]
    _header_row(ws, cols, row=6)

    AMT_COLS = {4, 5, 6}
    total_billed = 0.0
    total_paid   = 0.0
    total_due    = 0.0
    cur_row      = 7

    if party.opening_balance:
        ob_type = party.opening_balance_type or "debit"
        ob_due  = party.opening_balance if ob_type == "debit" else -party.opening_balance
        row = [
            "OPENING-BAL",
            party.created_at.strftime("%d-%m-%Y") if party.created_at else "—",
            "BALANCE",
            party.opening_balance,
            0.0,
            ob_due,
            "OPEN",
            "—",
            f"Opening balance ({ob_type})",
        ]
        _data_row(ws, row, cur_row, AMT_COLS)
        total_due += ob_due
        cur_row += 1

    unlinked_payments = payment_map.get(None, [])
    for up in unlinked_payments:
        row = [
            "SETTLEMENT",
            up.event_date.strftime("%d-%m-%Y"),
            "SETTLE-OB",
            0.0,
            up.amount,
            -up.amount,
            "PAID",
            "—",
            f"{up.event_date.strftime('%d-%m-%Y')} ₹{up.amount:.2f} ({up.mode}) {up.reference_no or ''}".strip(),
        ]
        _data_row(ws, row, cur_row, AMT_COLS)
        total_paid += up.amount
        total_due -= up.amount
        cur_row += 1

    for inv in invoices:
        plist = payment_map.get(inv.id, [])
        ptext = "; ".join(
            f"{p.event_date.strftime('%d-%m-%Y')} ₹{p.amount:.2f} ({p.mode})"
            for p in plist
        ) or "—"
        row = [
            inv.invoice_number,
            inv.invoice_date.strftime("%d-%m-%Y"),
            inv.invoice_type.upper(),
            inv.grand_total,
            inv.amount_paid,
            inv.amount_due,
            inv.payment_status.upper(),
            inv.gst_status.replace("_", " ").upper(),
            ptext,
        ]
        _data_row(ws, row, cur_row, AMT_COLS)
        total_billed += inv.grand_total
        total_paid   += inv.amount_paid
        total_due    += inv.amount_due
        cur_row += 1

    t_row = cur_row
    _total_row(ws, [
        "TOTAL", f"{len(invoices)} bills", "",
        round(total_billed, 2), round(total_paid, 2), round(total_due, 2),
        "", "", "",
    ], t_row, AMT_COLS)

    fname = f"ledger_{party.name[:20].replace(' ', '_')}_{date.today()}.xlsx"
    return _stream(wb, fname)


# ── 4. OLD GOLD REGISTER ──────────────────────────────────────────────────────

@router.get("/old-gold")
def export_old_gold(
    metal: str = "",
    session: Session = Depends(get_session),
):
    stmt = select(OldGoldExchange).order_by(OldGoldExchange.exchange_date)
    if metal in ["gold", "silver"]:
        stmt = stmt.where(OldGoldExchange.metal_type == metal)

    records = session.exec(stmt).all()

    party_ids = {r.party_id for r in records}
    parties   = {p.id: p for p in session.exec(
        select(Party).where(Party.id.in_(party_ids))
    ).all()} if party_ids else {}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Old Gold Register"
    ws.row_dimensions[1].height = 30

    cols = [
        ("Date",          12), ("Party Name",     22), ("Metal",          8),
        ("Type",          14), ("Purity",          8), ("Weight (g)",     12),
        ("Rate/g (₹)",    12), ("Total Value (₹)", 16), ("Cash Paid (₹)",  14),
        ("Notes",         25),
    ]
    _header_row(ws, cols, row=1)

    AMT_COLS = {6, 7, 8, 9}
    total_weight = 0.0
    total_value  = 0.0

    for r_idx, rec in enumerate(records, start=2):
        party = parties.get(rec.party_id)
        row = [
            rec.exchange_date.strftime("%d-%m-%Y"),
            party.name if party else "—",
            rec.metal_type.upper(),
            rec.transaction_type.replace("_", " ").upper(),
            rec.purity or "—",
            rec.weight_grams,
            rec.rate_per_gram,
            rec.total_value,
            rec.cash_paid or 0,
            rec.notes or "",
        ]
        _data_row(ws, row, r_idx, AMT_COLS)
        total_weight += rec.weight_grams
        total_value  += rec.total_value

    t_row = len(records) + 2
    _total_row(ws, [
        "TOTAL", f"{len(records)} records", "", "", "",
        round(total_weight, 3), "", round(total_value, 2), "", "",
    ], t_row, {6, 8})

    fname = f"old_gold_register_{metal or 'all'}_{date.today()}.xlsx"
    return _stream(wb, fname)


# ── 5. STOCK REGISTER ─────────────────────────────────────────────────────────

@router.get("/stock")
def export_stock(session: Session = Depends(get_session)):
    products = session.exec(
        select(Product).where(Product.is_active == True).order_by(Product.name)
    ).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Stock Register"
    ws.row_dimensions[1].height = 30

    cols = [
        ("Product Name",   24), ("Purity",         10), ("Metal",          10),
        ("HSN Code",       12), ("Total In (g)",    14), ("Total Out (g)",  14),
        ("Balance (g)",    14), ("Low Stock Alert", 16), ("Status",         12),
    ]
    _header_row(ws, cols, row=1)

    for r_idx, product in enumerate(products, start=2):
        entries = session.exec(
            select(StockLedger).where(StockLedger.product_id == product.id)
        ).all()
        if not entries:
            continue

        total_in  = round(sum(e.quantity_in  for e in entries), 3)
        total_out = round(sum(e.quantity_out for e in entries), 3)
        balance   = round(total_in - total_out, 3)
        is_low    = product.low_stock_alert is not None and balance <= product.low_stock_alert

        row = [
            product.name,
            product.purity or "—",
            product.metal_type.upper(),
            product.hsn_code,
            total_in,
            total_out,
            balance,
            product.low_stock_alert if product.low_stock_alert is not None else "—",
            "⚠ LOW" if is_low else "OK",
        ]
        _data_row(ws, row, r_idx, {5, 6, 7})

    fname = f"stock_register_{date.today()}.xlsx"
    return _stream(wb, fname)


# ── 6. EXPENSE REGISTER ───────────────────────────────────────────────────────

@router.get("/expenses")
def export_expenses(
    month: str = "",
    session: Session = Depends(get_session),
):
    stmt = (
        select(Expense)
        .where(Expense.is_deleted == False)
        .order_by(Expense.expense_date)
    )
    if month:
        try:
            y, m = int(month[:4]), int(month[5:7])
            last = calendar.monthrange(y, m)[1]
            stmt = stmt.where(Expense.expense_date >= date(y, m, 1))
            stmt = stmt.where(Expense.expense_date <= date(y, m, last))
        except Exception:
            pass

    expenses = session.exec(stmt).all()

    cat_ids    = {e.category_id for e in expenses}
    categories = {c.id: c for c in session.exec(
        select(ExpenseCategory).where(ExpenseCategory.id.in_(cat_ids))
    ).all()} if cat_ids else {}

    party_ids = {e.party_id for e in expenses if e.party_id}
    parties   = {p.id: p for p in session.exec(
        select(Party).where(Party.id.in_(party_ids))
    ).all()} if party_ids else {}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Expense Register"
    ws.row_dimensions[1].height = 30

    cols = [
        ("Date",           12), ("Category",       18), ("ITC Eligible",   12),
        ("Description",    26), ("Party",          20), ("Amount (₹)",      14),
        ("GST Paid (₹)",   14), ("ITC Claimable (₹)",16), ("Payment Mode",  14),
        ("Reference No.",  16),
    ]
    _header_row(ws, cols, row=1)

    AMT_COLS = {6, 7, 8}
    t_amount = t_gst = t_itc = 0.0

    for r_idx, exp in enumerate(expenses, start=2):
        cat   = categories.get(exp.category_id)
        party = parties.get(exp.party_id) if exp.party_id else None
        row = [
            exp.expense_date.strftime("%d-%m-%Y"),
            cat.name if cat else "—",
            "YES" if (cat and cat.is_itc_eligible) else "NO",
            exp.description or "—",
            party.name if party else "—",
            exp.amount,
            exp.gst_amount or 0,
            exp.itc_claimable or 0,
            exp.payment_mode.upper() if exp.payment_mode else "—",
            exp.reference_no or "—",
        ]
        _data_row(ws, row, r_idx, AMT_COLS)
        t_amount += exp.amount
        t_gst    += exp.gst_amount or 0
        t_itc    += exp.itc_claimable or 0

    t_row = len(expenses) + 2
    _total_row(ws, [
        "TOTAL", f"{len(expenses)} expenses", "", "", "",
        round(t_amount, 2), round(t_gst, 2), round(t_itc, 2), "", "",
    ], t_row, AMT_COLS)

    fname = f"expenses_{month or 'all'}_{date.today()}.xlsx"
    return _stream(wb, fname)